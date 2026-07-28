"""Generador del presupuesto en Excel — el entregable real de la papelería.

QUÉ PROBLEMA RESUELVE. El modelo NO puede generar este archivo: `ToolPolicy.
companion()` le bloquea `Bash`, `Write` y `Edit` (og118 #277 — un companion no
debe poder tocar el filesystem del host). Verificado en runtime el 27-jul: ante
"genera el Excel" el modelo intenta `Bash`, falla, y termina entregando una
tabla markdown en el chat. Parece útil y no lo es: lo que se manda por WhatsApp
a la mamá del alumno es un .xlsx, no un mensaje.

LA SOLUCIÓN, y por qué es mejor que darle Bash. El modelo produce los DATOS y el
servidor aplica el FORMATO. Dos ganancias en un movimiento:

1. Seguridad: no hace falta abrirle ejecución de código arbitrario a un
   companion para producir una hoja de cálculo.
2. El formato queda invariante por construcción. Las Instructions repiten
   "PLANTILLA … NO alterar formato" precisamente porque un modelo que ejecuta el
   script puede improvisar el diseño. Si el script vive aquí, no hay nada que
   improvisar.

El diseño es réplica fiel del presupuesto aprobado por la dirección (header
negro, columna "Precio c/desc." por renglón, banda de forrado, total en negro),
tomado de `muestrario-excels/plantilla_fenix.py`.
"""

from __future__ import annotations

import datetime as _dt
import io
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NEGRO = "000000"
GRIS_BANDA = "808080"
GRIS_HDR = "BFBFBF"
BLANCO = "FFFFFF"
GRIS_LINEA = "BFBFBF"
MONEDA = '"$"#,##0.00'

# Datos de contacto canónicos (confirmados por la dirección el 20-jul). El
# muestrario traía "Tel. 33 458226", que estaba truncado.
DIRECCION = "José María Gómez #476, San Juan Bosco"
CONTACTO = "WhatsApp 33 3858 2967   |   Tel. 33 3345 8226"

_thin = Side(style="thin", color=GRIS_LINEA)
BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


@dataclass
class Renglon:
    descripcion: str
    cantidad: float
    precio: float  # precio de LISTA; el descuento lo aplica la hoja


@dataclass
class Presupuesto:
    alumno: str = ""
    escuela: str = ""
    grado: str = ""
    tutor: str = ""
    fecha: str = ""
    descuento: float = 0.15
    items: list[Renglon] = field(default_factory=list)
    forrado: list[Renglon] = field(default_factory=list)
    # Se MUESTRAN con su precio pero NO suman al total: el cliente decide.
    opcionales: list[Renglon] = field(default_factory=list)
    # Lo que Fénix no maneja: sin precio, para que el cliente sepa que lo
    # consigue en otro lado en vez de creer que su lista estaba completa.
    fuera: list[str] = field(default_factory=list)


def _relleno(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def generar(p: Presupuesto) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    # Los stubs de openpyxl declaran `active` como Optional; en un Workbook
    # recién creado nunca lo es.
    assert ws is not None
    ws.title = "Presupuesto"
    ws.sheet_view.showGridLines = False
    for col, ancho in {"A": 5, "B": 50, "C": 7, "D": 12, "E": 13, "F": 13}.items():
        ws.column_dimensions[col].width = ancho

    def centrado(rango: str, texto: str, **fuente):
        ws.merge_cells(rango)
        c = ws[rango.split(":")[0]]
        c.value = texto
        c.font = Font(**fuente)
        c.alignment = Alignment(horizontal="center", vertical="center")
        return c

    centrado("A1:F1", "SERVICIOS PAPELEROS FÉNIX", name="Arial", size=18, bold=True, color=BLANCO)
    centrado("A2:F2", f"{DIRECCION}   |   {CONTACTO}", name="Arial", size=9.5, color=BLANCO)
    for fila_ in (1, 2):
        for col in range(1, 7):
            ws.cell(fila_, col).fill = _relleno(NEGRO)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 16

    hoy = _dt.date.today()
    fecha = p.fecha or f"{hoy.day:02d}/{MESES[hoy.month - 1]}/{hoy.year}"
    escuela = " — ".join(x for x in (p.escuela, p.grado) if x) or "Lista de útiles"
    pct = round(p.descuento * 100)

    centrado("A4:F4", "PRESUPUESTO — LISTA DE ÚTILES ESCOLARES", size=13, bold=True, color=NEGRO)
    centrado("A5:F5", escuela, size=11, italic=True, color="404040")
    linea_alumno = f"Alumno: {p.alumno or '—'}" + (f"        Mamá: {p.tutor}" if p.tutor else "")
    centrado("A6:F6", linea_alumno, size=11, bold=True, color=NEGRO)
    centrado(
        "A7:F7",
        f"Fecha: {fecha}        Promoción temporada escolar: {pct}% de descuento",
        size=10,
        color="404040",
    )
    for fila_ in (4, 5, 6, 7):
        ws.row_dimensions[fila_].height = 18

    encabezado = 9
    for i, titulo in enumerate(
        ["#", "Descripción", "Cant.", "P. Unitario", "Precio c/desc.", "Importe"], 1
    ):
        c = ws.cell(encabezado, i, titulo)
        c.fill = _relleno(NEGRO)
        c.font = Font(size=10, bold=True, color=BLANCO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDE
    ws.row_dimensions[encabezado].height = 22

    estado = {"fila": encabezado + 1, "n": 1, "subtotal": 0.0}

    def renglon(r: Renglon, zebra: bool, con_descuento: bool = True, suma: bool = True) -> None:
        f = estado["fila"]
        cd = round(r.precio * (1 - p.descuento), 2) if con_descuento else round(r.precio, 2)
        importe = round(r.cantidad * cd, 2)
        ws.cell(f, 1, estado["n"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(f, 2, r.descripcion).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )
        ws.cell(f, 3, r.cantidad).alignment = Alignment(horizontal="center", vertical="center")
        for col, valor in ((4, r.precio), (5, cd), (6, importe)):
            celda = ws.cell(f, col, valor)
            celda.number_format = MONEDA
            celda.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        for col in range(1, 7):
            ws.cell(f, col).font = Font(size=10)
            ws.cell(f, col).fill = _relleno(GRIS_HDR if zebra else BLANCO)
            ws.cell(f, col).border = BORDE
        ws.row_dimensions[f].height = 15
        if suma:
            estado["subtotal"] += r.cantidad * r.precio
        estado["fila"] += 1
        estado["n"] += 1

    def banda(texto: str, color: str = GRIS_BANDA) -> None:
        f = estado["fila"]
        ws.merge_cells(f"A{f}:F{f}")
        ws.cell(f, 1, texto).font = Font(size=10, bold=True, color=BLANCO)
        ws.cell(f, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for col in range(1, 7):
            ws.cell(f, col).fill = _relleno(color)
        ws.row_dimensions[f].height = 18
        estado["fila"] += 1

    for i, r in enumerate(p.items):
        renglon(r, i % 2 == 1)

    if p.forrado:
        banda("SERVICIO DE FORRADO")
        for i, r in enumerate(p.forrado):
            renglon(r, i % 2 == 1)

    subtotal = round(estado["subtotal"], 2)
    ahorro = round(subtotal * p.descuento, 2)
    total = round(subtotal - ahorro, 2)

    def total_fila(etiqueta: str, valor: float, negro: bool = False) -> None:
        f = estado["fila"]
        ws.merge_cells(f"A{f}:E{f}")
        lc = ws.cell(f, 1, etiqueta)
        lc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        vc = ws.cell(f, 6, valor)
        vc.number_format = MONEDA
        vc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        if negro:
            for col in range(1, 7):
                ws.cell(f, col).fill = _relleno(NEGRO)
            lc.font = Font(size=12, bold=True, color=BLANCO)
            vc.font = Font(size=12, bold=True, color=BLANCO)
            ws.row_dimensions[f].height = 22
        else:
            lc.font = Font(size=10, bold=True)
            vc.font = Font(size=10, bold=True)
            ws.row_dimensions[f].height = 16
        lc.border = BORDE
        vc.border = BORDE
        estado["fila"] += 1

    total_fila("Subtotal sin descuento:", subtotal)
    total_fila(f"Ahorro ({pct}% descuento):", ahorro)
    total_fila("TOTAL A PAGAR:", total, negro=True)
    estado["fila"] += 1

    if p.opcionales:
        # Banda dorada: se ven con su costo pero NO están sumadas arriba. La
        # rotulación va a precio de lista — regla fija de la dirección.
        banda("OPCIONALES — no incluidos en el total", "BF8F00")
        for i, r in enumerate(p.opcionales):
            renglon(r, i % 2 == 1, con_descuento=False, suma=False)
        estado["fila"] += 1

    if p.fuera:
        banda("FUERA DEL PRESUPUESTO — no lo manejamos en Fénix", "A6A6A6")
        for texto in p.fuera:
            f = estado["fila"]
            ws.merge_cells(f"A{f}:F{f}")
            c = ws.cell(f, 1, texto)
            c.font = Font(size=10, italic=True, color="404040")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            ws.row_dimensions[f].height = 15
            estado["fila"] += 1
        estado["fila"] += 1

    centrado(
        f"A{estado['fila']}:F{estado['fila']}",
        "Precios sujetos a existencia. Presupuesto válido durante la temporada escolar."
        "   ***  Gracias por su compra  ***",
        size=9,
        italic=True,
        color="808080",
    )

    ws.freeze_panes = f"A{encabezado + 1}"
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nombre_archivo(p: Presupuesto) -> str:
    """`Presupuesto-Emma-Hernandez-4B.xlsx` — se manda por WhatsApp, así que el
    nombre tiene que decir de quién es sin abrirlo."""
    import re
    import unicodedata

    partes = [x for x in (p.alumno or "cliente", p.grado) if x]
    base = " ".join(partes)
    base = unicodedata.normalize("NFKD", base).encode("ascii", "ignore").decode()
    base = re.sub(r"[^A-Za-z0-9]+", "-", base).strip("-") or "cliente"
    return f"Presupuesto-{base}.xlsx"


def a_vista(datos: bytes) -> dict:
    """El .xlsx generado → JSON para el visor del navegador.

    Se parsea EL ARCHIVO, no los datos de entrada. Si el visor se dibujara desde
    el input, podría mostrar algo distinto a lo que se descarga — y una vista
    previa que no es fiel al archivo es peor que ninguna, porque se confía en
    ella. Una sola fuente: el archivo.
    """
    import openpyxl

    ws = openpyxl.load_workbook(io.BytesIO(datos)).active
    assert ws is not None

    # Rangos combinados, para que el visor pinte las bandas como en Excel.
    combinadas = {}
    for rango in ws.merged_cells.ranges:
        combinadas[(rango.min_row, rango.min_col)] = (
            rango.max_col - rango.min_col + 1,
            rango.max_row - rango.min_row + 1,
        )
    ocultas = {
        (f, c)
        for rango in ws.merged_cells.ranges
        for f in range(rango.min_row, rango.max_row + 1)
        for c in range(rango.min_col, rango.max_col + 1)
        if (f, c) != (rango.min_row, rango.min_col)
    }

    filas = []
    for f in range(1, ws.max_row + 1):
        celdas = []
        for c in range(1, 7):
            if (f, c) in ocultas:
                continue
            celda = ws.cell(f, c)
            fuente = celda.font
            relleno = celda.fill
            color_fondo = None
            if relleno is not None and relleno.fgColor is not None and relleno.patternType:
                rgb = relleno.fgColor.rgb
                if isinstance(rgb, str) and len(rgb) == 8:
                    color_fondo = f"#{rgb[2:]}"
            color_texto = None
            if fuente is not None and fuente.color is not None:
                rgb = fuente.color.rgb
                if isinstance(rgb, str) and len(rgb) == 8:
                    color_texto = f"#{rgb[2:]}"
            ancho, alto = combinadas.get((f, c), (1, 1))
            valor = celda.value
            celdas.append(
                {
                    "v": "" if valor is None else valor,
                    "moneda": bool(celda.number_format and "$" in celda.number_format),
                    "negrita": bool(fuente and fuente.bold),
                    "cursiva": bool(fuente and fuente.italic),
                    "tam": float(fuente.size) if fuente and fuente.size else 10.0,
                    "fondo": color_fondo,
                    "color": color_texto,
                    "alineado": (celda.alignment.horizontal if celda.alignment else None) or "left",
                    "cols": ancho,
                    "filas": alto,
                }
            )
        if celdas:
            filas.append(celdas)

    anchos = [ws.column_dimensions[c].width or 10 for c in "ABCDEF"]
    return {"filas": filas, "anchos": anchos}
